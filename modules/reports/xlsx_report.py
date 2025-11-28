"""
Excel Report Generator

Generates Excel weather reports with formatted data and charts.
Uses openpyxl for Excel generation.
"""

import io
from typing import Dict, Any, List, Optional
from datetime import datetime, timezone

from config import settings
from logging_config import get_logger

logger = get_logger(__name__)


class ExcelReportGenerator:
    """
    Generates Excel weather reports.
    
    Features:
    - Formatted weather data sheets
    - Multiple sheets for different data views
    - Basic styling and headers
    """
    
    def __init__(self):
        """Initialize the Excel generator."""
        self._branding = {
            "title": "IntelliWeather Report",
            "company": "IntelliWeather",
            "tagline": "Your Premium Weather Companion"
        }
    
    def generate(
        self,
        location: str,
        weather_data: Dict[str, Any],
        report_type: str = "daily"
    ) -> bytes:
        """
        Generate an Excel report.
        
        Args:
            location: Location name
            weather_data: Weather data dict
            report_type: "hourly" or "daily"
            
        Returns:
            Excel content as bytes
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
            
            return self._generate_with_openpyxl(location, weather_data, report_type)
        except ImportError:
            logger.warning("openpyxl not available, generating CSV instead")
            return self._generate_csv_fallback(location, weather_data, report_type)
    
    def _generate_with_openpyxl(
        self,
        location: str,
        weather_data: Dict[str, Any],
        report_type: str
    ) -> bytes:
        """Generate Excel using openpyxl."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Weather Report"
        
        # Styles
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True, size=16, color="4A1C6E")
        cell_font = Font(size=11)
        
        header_fill = PatternFill(start_color="4A1C6E", end_color="4A1C6E", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title Section
        ws['A1'] = self._branding['title']
        ws['A1'].font = title_font
        ws.merge_cells('A1:F1')
        
        ws['A2'] = f"Location: {location}"
        ws['A2'].font = Font(size=12)
        
        ws['A3'] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}"
        ws['A3'].font = Font(size=10, italic=True)
        
        ws['A4'] = f"Report Type: {report_type.title()} Forecast"
        ws['A4'].font = Font(size=10)
        
        # Data section starts at row 6
        start_row = 6
        
        if report_type == "hourly":
            headers = ["Time", "Temperature (°C)", "Humidity (%)", "Wind (m/s)", "Precip. Prob. (%)", "Weather"]
            data = weather_data.get("hourly", [])
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font_white
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            
            # Write data
            for row_idx, hour in enumerate(data[:48], start_row + 1):
                ws.cell(row=row_idx, column=1, value=hour.get("time", "")[:16]).border = thin_border
                ws.cell(row=row_idx, column=2, value=hour.get("temperature")).border = thin_border
                ws.cell(row=row_idx, column=3, value=hour.get("humidity")).border = thin_border
                ws.cell(row=row_idx, column=4, value=hour.get("wind_speed")).border = thin_border
                ws.cell(row=row_idx, column=5, value=hour.get("precipitation_probability")).border = thin_border
                ws.cell(row=row_idx, column=6, value=self._get_condition(hour.get("weather_code", 0))).border = thin_border
        
        else:  # daily
            headers = ["Date", "Max Temp (°C)", "Min Temp (°C)", "Precipitation (mm)", "Precip. Prob. (%)", "Sunrise", "Sunset", "Weather"]
            data = weather_data.get("daily", [])
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font_white
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            
            # Write data
            for row_idx, day in enumerate(data, start_row + 1):
                ws.cell(row=row_idx, column=1, value=day.get("date")).border = thin_border
                ws.cell(row=row_idx, column=2, value=day.get("temperature_max")).border = thin_border
                ws.cell(row=row_idx, column=3, value=day.get("temperature_min")).border = thin_border
                ws.cell(row=row_idx, column=4, value=day.get("precipitation_sum")).border = thin_border
                ws.cell(row=row_idx, column=5, value=day.get("precipitation_probability_max")).border = thin_border
                ws.cell(row=row_idx, column=6, value=day.get("sunrise", "")[-5:] if day.get("sunrise") else "").border = thin_border
                ws.cell(row=row_idx, column=7, value=day.get("sunset", "")[-5:] if day.get("sunset") else "").border = thin_border
                ws.cell(row=row_idx, column=8, value=self._get_condition(day.get("weather_code", 0))).border = thin_border
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Add metadata sheet
        ws_meta = wb.create_sheet("Metadata")
        ws_meta['A1'] = "IntelliWeather Report Metadata"
        ws_meta['A1'].font = title_font
        
        ws_meta['A3'] = "Data Source:"
        ws_meta['B3'] = "Open-Meteo API"
        
        ws_meta['A4'] = "Report Generated:"
        ws_meta['B4'] = datetime.now(timezone.utc).isoformat()
        
        ws_meta['A5'] = "Location:"
        ws_meta['B5'] = location
        
        ws_meta['A6'] = "Latitude:"
        ws_meta['B6'] = weather_data.get("latitude", "N/A")
        
        ws_meta['A7'] = "Longitude:"
        ws_meta['B7'] = weather_data.get("longitude", "N/A")
        
        ws_meta['A8'] = "Timezone:"
        ws_meta['B8'] = weather_data.get("timezone", "N/A")
        
        ws_meta['A10'] = "Powered by IntelliWeather"
        ws_meta['A10'].font = Font(italic=True)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def _generate_csv_fallback(
        self,
        location: str,
        weather_data: Dict[str, Any],
        report_type: str
    ) -> bytes:
        """Generate CSV as fallback when openpyxl is not available."""
        import csv
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([f"IntelliWeather Report - {location}"])
        writer.writerow([f"Generated: {datetime.now(timezone.utc).isoformat()}"])
        writer.writerow([])
        
        if report_type == "hourly":
            writer.writerow(["Time", "Temperature (°C)", "Humidity (%)", "Wind (m/s)", "Precip. Prob. (%)"])
            for hour in weather_data.get("hourly", [])[:48]:
                writer.writerow([
                    hour.get("time", ""),
                    hour.get("temperature"),
                    hour.get("humidity"),
                    hour.get("wind_speed"),
                    hour.get("precipitation_probability")
                ])
        else:
            writer.writerow(["Date", "Max Temp (°C)", "Min Temp (°C)", "Precipitation (mm)", "Weather"])
            for day in weather_data.get("daily", []):
                writer.writerow([
                    day.get("date"),
                    day.get("temperature_max"),
                    day.get("temperature_min"),
                    day.get("precipitation_sum"),
                    self._get_condition(day.get("weather_code", 0))
                ])
        
        return output.getvalue().encode('utf-8')
    
    def _get_condition(self, code: int) -> str:
        """Get text description for weather code."""
        conditions = {
            0: "Clear",
            1: "Mostly Clear",
            2: "Partly Cloudy",
            3: "Overcast",
            45: "Foggy",
            48: "Fog",
            51: "Light Drizzle",
            53: "Drizzle",
            55: "Heavy Drizzle",
            61: "Light Rain",
            63: "Rain",
            65: "Heavy Rain",
            71: "Light Snow",
            73: "Snow",
            75: "Heavy Snow",
            80: "Rain Showers",
            81: "Heavy Showers",
            82: "Violent Showers",
            95: "Thunderstorm",
            96: "T-Storm + Hail",
            99: "Severe T-Storm"
        }
        return conditions.get(code, "Unknown")
